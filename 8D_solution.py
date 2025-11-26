
import streamlit as st
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io

# --- Styles ---
title_fill = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
section_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
thin = Side(style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

st.title("8D Problem-Solving Report Generator (Ishikawa auto-embedded)")

# --- Header Info ---
st.header("Header Information")
doc_no = st.text_input("Document No.", "FM2-011a", key="doc_no")
change_no = st.text_input("Change No.", "23-022", key="change_no")
issue_date = st.date_input("Issue Date", key="issue_date")
rev_no = st.text_input("Rev. No", "3", key="rev_no")
product_name = st.text_input("Product Name", key="product_name")
rma_no = st.text_input("RMA No", key="rma_no")
product_model = st.text_input("Product Model", key="product_model")
received_date = st.date_input("Received Date", key="received_date")
notification_date = st.date_input("Notification Date", key="notification_date")
serial_imei = st.text_input("Serial Number / IMEI", key="serial_imei")

# --- Sections ---
st.header("1 - Problem Description")
problem_desc = st.text_area("Problem Description", key="problem_desc")

st.header("2 - Team Members")
team_count = st.number_input("Number of Team Members", 1, 10, 4, key="team_count")
team_members = []
for i in range(team_count):
    name = st.text_input(f"Name {i+1}", key=f"team_name_{i}")
    dept = st.text_input(f"Department {i+1}", key=f"team_dept_{i}")
    team_members.append((name, dept))

st.header("3 - Containment Actions")
containment_count = st.number_input("Number of Containment Actions", 1, 10, 2, key="containment_count")
containment_actions = []
for i in range(containment_count):
    action = st.text_input(f"Containment Action {i+1}", key=f"contain_action_{i}")
    responsible = st.text_input(f"Responsible {i+1}", key=f"contain_resp_{i}")
    date = st.date_input(f"Date {i+1}", key=f"contain_date_{i}")
    containment_actions.append({"action": action, "responsible": responsible, "date": str(date)})

st.header("4 - Investigation")
what = st.text_area("WHAT", key="what")
how = st.text_area("HOW", key="how")
who = st.text_area("WHO", key="who")
where = st.text_area("WHERE", key="where")

# --- Ishikawa Inputs (auto used during Excel generation) ---
st.subheader("Ishikawa Diagram Inputs (auto-generated with report)")
machine = st.text_area("Machine", key="machine")
method = st.text_area("Method", key="method")
material = st.text_area("Material", key="material")
manpower = st.text_area("Manpower", key="manpower")
measurement = st.text_area("Measurement", key="measurement")
environment = st.text_area("Environment", key="environment")

def parse_lines(txt: str):
    return [line.strip() for line in txt.splitlines() if line.strip()]

# --- Improved Ishikawa Diagram Function (auto-generated during Excel build) ---
def build_ishikawa(categories, problem_title="Problem"):
    """
    Creates a Fishbone diagram image and returns a PNG BytesIO buffer.
    categories: dict with keys: Machine, Method, Material, Manpower, Measurement, Environment
               values are lists of sub-causes (strings).
    """
    fig, ax = plt.subplots(figsize=(12, 8))  # Larger canvas for spacing
    ax.set_title("Ishikawa (Fishbone) Diagram", fontsize=16, fontweight="bold")
    ax.axis("off")

    # Main spine
    ax.plot([0.05, 0.95], [0.5, 0.5], color="black", linewidth=2)
    ax.text(0.96, 0.5, problem_title, fontsize=14, va="center", fontweight="bold")

    # Category positions (spread out horizontally)
    cat_order = ["Machine", "Method", "Material", "Manpower", "Measurement", "Environment"]
    top_cats = cat_order[:3]
    bottom_cats = cat_order[3:]
    x_positions = [0.15, 0.35, 0.55, 0.15, 0.35, 0.55]

    # Top branches
    for idx, cat in enumerate(top_cats):
        x = x_positions[idx]
        ax.plot([x, 0.5], [0.5, 0.7], color="black", linewidth=1.5)
        ax.text(x - 0.02, 0.72, cat, fontsize=12, ha="right", fontweight="bold")
        causes = categories.get(cat, [])
        for j, cause in enumerate(causes[:8]):  # allow up to 8 sub-causes with spaced ticks
            y = 0.72 + 0.035 * (j + 1)  # vertical spacing
            ax.plot([x + 0.02, x + 0.09], [y, y], color="black")
            ax.text(x + 0.10, y, f"- {cause}", fontsize=10, va="center")

    # Bottom branches
    for idx, cat in enumerate(bottom_cats):
        x = x_positions[idx + 3]
        ax.plot([x, 0.5], [0.5, 0.3], color="black", linewidth=1.5)
        ax.text(x - 0.02, 0.28, cat, fontsize=12, ha="right", fontweight="bold")
        causes = categories.get(cat, [])
        for j, cause in enumerate(causes[:8]):
            y = 0.28 - 0.035 * (j + 1)  # vertical spacing downward
            ax.plot([x + 0.02, x + 0.09], [y, y], color="black")
            ax.text(x + 0.10, y, f"- {cause}", fontsize=10, va="center")

    buf = BytesIO()
    plt.tight_layout()
    fig.savefig(buf, format="png", dpi=160, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf

# --- Remaining Sections ---
st.header("5 - Root Cause")
root_cause = st.text_area("Root Cause Description", key="root_cause")

st.header("6 - Corrective Actions")
corrective_count = st.number_input("Number of Corrective Actions", 1, 10, 2, key="corrective_count")
corrective_actions = []
for i in range(corrective_count):
    action = st.text_input(f"Corrective Action {i+1}", key=f"corr_action_{i}")
    responsible = st.text_input(f"Responsible {i+1}", key=f"corr_resp_{i}")
    date = st.date_input(f"Date {i+1}", key=f"corr_date_{i}")
    corrective_actions.append({"action": action, "responsible": responsible, "date": str(date)})

st.header("7 - Preventive Actions")
preventive_count = st.number_input("Number of Preventive Actions", 1, 10, 2, key="preventive_count")
preventive_actions = []
for i in range(preventive_count):
    action = st.text_input(f"Preventive Action {i+1}", key=f"prev_action_{i}")
    responsible = st.text_input(f"Responsible {i+1}", key=f"prev_resp_{i}")
    date = st.date_input(f"Date {i+1}", key=f"prev_date_{i}")
    preventive_actions.append({"action": action, "responsible": responsible, "date": str(date)})

# --- Generate Excel (Ishikawa included automatically) ---
if st.button("Generate 8D Report"):
    # Build Ishikawa data (always, no separate button)
    categories = {
        "Machine": parse_lines(machine),
        "Method": parse_lines(method),
        "Material": parse_lines(material),
        "Manpower": parse_lines(manpower),
        "Measurement": parse_lines(measurement),
        "Environment": parse_lines(environment),
    }
    # Generate the diagram image buffer
    ishikawa_buf = build_ishikawa(categories, problem_title="Problem")

    # Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "8D"
    for i, w in enumerate([18,18,22,12,20,18,18,18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value="8D PROBLEM SOLUTION REPORT").font = Font(size=16,bold=True)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=1).fill = title_fill
    row += 2

    for label, value in [("DOCUMENT NO.", doc_no),("CHANGE NO.", change_no),("ISSUE DATE", str(issue_date)),("REV. NO", rev_no)]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=3, value=value)
        for c in range(1,4): ws.cell(row=row, column=c).border = border
        row += 1
    row += 1

    def add_kv(label, value):
        global row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=1, value=f"{label}:").font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
        ws.cell(row=row, column=3, value=value)
        for c in range(1,9): ws.cell(row=row, column=c).border = border
        row += 1

    for lab,val in [("Product Name",product_name),("RMA No",rma_no),("Product Model",product_model),("Received Date",str(received_date)),("Serial Number/ IMEI",serial_imei),("Notification Date",str(notification_date))]:
        add_kv(lab,val)
    row += 1

    def section(title):
        global row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row=row, column=1, value=title).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = section_fill
        row += 1

    section("1- PROBLEM DESCRIPTION")
    ws.merge_cells(start_row=row, start_column=1, end_row=row+2, end_column=8)
    ws.cell(row=row, column=1, value=problem_desc)
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    for rr in range(row, row+3):
        for c in range(1, 9): ws.cell(row=rr, column=c).border = border
    row += 4

    section("2- TEAM MEMBERS")
    ws.cell(row=row, column=1, value="NAME").font = Font(bold=True)
    ws.cell(row=row, column=2, value="DEPARTMENT").font = Font(bold=True)
    for c in range(1, 3): ws.cell(row=row, column=c).border = border
    row += 1
    for name, dept in team_members:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=dept)
        for c in range(1, 3): ws.cell(row=row, column=c).border = border
        row += 1
    row += 1

    def actions_section(title, items):
        global row
        section(title)
        ws.cell(row=row, column=1, value="ACTION").font = Font(bold=True)
        ws.cell(row=row, column=6, value="RESPONSIBLE").font = Font(bold=True)
        ws.cell(row=row, column=8, value="DATE").font = Font(bold=True)
        for c in [1, 6, 8]: ws.cell(row=row, column=c).border = border
        row += 1
        for it in items:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            ws.cell(row=row, column=1, value=it['action'])
            ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
            ws.cell(row=row, column=6, value=it['responsible'])
            ws.cell(row=row, column=8, value=it['date'])
            for c in range(1, 9): ws.cell(row=row, column=c).border = border
            row += 1
        row += 1

    actions_section("3- CONTAINMENT ACTIONS", containment_actions)

    section("4- INVESTIGATION")
    for lab, val in [("WHAT", what), ("HOW", how), ("WHO", who), ("WHERE", where)]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=1, value=f"{lab}:").font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
        ws.cell(row=row, column=3, value=val)
        for c in range(1, 9): ws.cell(row=row, column=c).border = border
        row += 1
    row += 1

    # Auto-embed Ishikawa diagram here (no separate button)
    img = Image(BytesIO(ishikawa_buf.getvalue()))
    img.width, img.height = 600, 400  # adjust size as needed
    ws.add_image(img, f"A{row}")
    row += 20  # advance rows to avoid overlap with next section

    section("5- ROOT CAUSE")
    ws.merge_cells(start_row=row, start_column=1, end_row=row+2, end_column=8)
    ws.cell(row=row, column=1, value=root_cause)
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    for rr in range(row, row+3):
        for c in range(1, 9): ws.cell(row=rr, column=c).border = border
    row += 4

    actions_section("6- CORRECTIVE ACTIONS", corrective_actions)
    actions_section("7- PREVENTIVE ACTIONS", preventive_actions)

    ws.cell(row=row, column=1, value="Made by:").font = Font(bold=True)
    ws.cell(row=row, column=4, value="Review by:").font = Font(bold=True)
    ws.cell(row=row, column=7, value="Approve By:").font = Font(bold=True)
    row += 1
    for start_col in [1, 4, 7]:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row+1, end_column=start_col+2)
        for rr in range(row, row+2):
            for cc in range(start_col, start_col+3):
                ws.cell(row=rr, column=cc).border = border

    # Save to BytesIO and offer download
    output = io.BytesIO()
    wb.save(output)
    st.download_button("Download Excel", data=output.getvalue(), file_name="8D_Report_with_Ishikawa.xlsx")
